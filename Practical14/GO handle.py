import xml.dom.minidom as minidom
from openpyxl import Workbook
# Parse the XML file
dom = minidom.parse("C:/Users/DELL/OneDrive/桌面/progaming/go_obo.xml")
root = dom.documentElement
terms = root.getElementsByTagName('term')

nodeparents = {}
result = {}
#we should write the relation between two nodes in a dic (key is the child and value are parents) because dic is suitable for checking
for term in terms:
    node_id = term.getElementsByTagName("id")[0].childNodes[0].data
    ancestor_ids = []
    for is_a in term.getElementsByTagName("is_a"):
        ancestor_ids.append(is_a.childNodes[0].data)
    nodeparents[node_id] = ancestor_ids
    result[node_id] = 0
    # Initialize the "reult" which is used to store the answer

def findparentsnodes(node_id, parent_ids):
    ancestor_ids = nodeparents[node_id]
    #take parent_ids as a set to store all parents relationship 
    for ancestor_id in ancestor_ids:
        parent_ids.add(ancestor_id) 
        findparentsnodes(ancestor_id, parent_ids)
        #parents_ids means all ancestor node (including parent's parents...etc)

#here is to get the final result part
for key in nodeparents.keys():
    parent_ids = set()
    # reset the parent_ids
    findparentsnodes(key, parent_ids)
    for parent_id in parent_ids:
        result[parent_id] += 1
        #if a node is included in another node's ancestor it surely should take it into the childnodes
        
# Initialize Excel workbook and sheet
workbook = Workbook()
sheet = workbook.active
sheet.title = 'GO Terms'

# Set column headers
headers = ['GO_ID', 'Term Name', 'Definition', 'Child Nodes']
for col_num, header in enumerate(headers, 1):
    sheet.cell(row=1, column=col_num, value=header)

# Find 'autophagosome' related gene ontology terms and their child nodes
row_num = 2  # Start from row 2        
for term in terms:
    go_id = term.getElementsByTagName('id')[0].childNodes[0].data
    term_name = term.getElementsByTagName('name')[0].childNodes[0].data
    defstr = term.getElementsByTagName('defstr')[0].childNodes[0].data

    # Check if 'autophagosome' is present in the definition string
    if 'autophagosome' in defstr.lower():
        child_count = result[go_id]
        # Write the information to the spreadsheet
        sheet.cell(row=row_num, column=1, value=go_id)
        sheet.cell(row=row_num, column=2, value=term_name)
        sheet.cell(row=row_num, column=3, value=defstr)
        sheet.cell(row=row_num, column=4, value=child_count)
        row_num += 1

# Save the workbook
workbook.save('C:/Users/DELL/OneDrive/桌面/progaming/go_terms.xlsx')

# I suppose there is many same counting process during the code running, especially in the findparentsnodes part. However, I didn't come up with a better one and maybe I can use something to store the things have already counted
#The result is shown in a another file
